from openpyxl import load_workbook

#수식 그대로 가져옴
# wb = load_workbook("sample_fomula.xlsx")
# ws = wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None이라 나옴 None 없앨려면 열어서 다시 저장하면 됨 
wb = load_workbook("sample_fomula.xlsx" , data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
         print(cell)